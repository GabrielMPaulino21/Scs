import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

# --- 1. CONFIGURAÇÃO DA PÁGINA ---
st.set_page_config(
    page_title="FollowUP-GY Automação",
    page_icon="🤖",
    layout="wide"
)

# --- 2. SUAS FUNÇÕES ORIGINAIS (COM MÍNIMAS ADAPTAÇÕES DE I/O) ---

def executar_planilhas_py(arquivo_cji5, arquivo_srm):
    """
    Contém a lógica EXATA do seu Planilhas.py.
    A única mudança é que lê os arquivos da memória e retorna o resultado em vez de salvar.
    """
    st.write("▶️ Etapa 1: Processando `Planilhas.py`...")
    
    # --- Início da sua lógica original ---
    df_cji5 = pd.read_excel(arquivo_cji5)
    df_srm = pd.read_excel(arquivo_srm)
    
    df_cji5['Nº doc.de referência'] = df_cji5['Nº doc.de referência'].astype(str)
    df_cji5 = df_cji5[df_cji5['Nº doc.de referência'].str.startswith('S', na=False)].copy()
    if df_cji5.empty: 
        st.warning("Etapa 1: Nenhuma SC encontrada no arquivo Cji5. O processo será interrompido.")
        return pd.DataFrame()
        
    df_cji5['SC_ID_Key'] = df_cji5['Nº doc.de referência'].str.replace('S', '', n=1, regex=False).str.strip()
    df_cji5['SC_ID_Key'] = pd.to_numeric(df_cji5['SC_ID_Key'], errors='coerce')
    df_cji5.dropna(subset=['SC_ID_Key'], inplace=True)
    df_cji5['SC_ID_Key'] = df_cji5['SC_ID_Key'].astype(int).astype(str)
    
    coluna_valor_correta = 'Valor/moed.transação'
    df_cji5[coluna_valor_correta] = pd.to_numeric(df_cji5[coluna_valor_correta], errors='coerce').fillna(0)
    
    agg_funcs = {
        'Material': lambda x: ';\n'.join(x.unique()),
        'Denominação': lambda x: ';\n'.join(x.unique()),
        'Quantidade total': lambda x: ';\n'.join(x),
        coluna_valor_correta: 'sum',
        'Nº doc.de referência': 'first'
    }
    df_agrupado = df_cji5.groupby(['Definição do projeto', 'SC_ID_Key']).agg(agg_funcs).reset_index()
    
    if 'SC ID' not in df_srm.columns:
        st.error("ERRO CRÍTICO: A coluna 'SC ID' não foi encontrada no arquivo DADOS_SRM.xlsx!")
        return None
    df_srm['SC_ID_Key'] = pd.to_numeric(df_srm['SC ID'], errors='coerce')
    df_srm.dropna(subset=['SC_ID_Key'], inplace=True)
    df_srm['SC_ID_Key'] = df_srm['SC_ID_Key'].astype(int).astype(str)
    df_srm = df_srm.drop_duplicates(subset=['SC_ID_Key'], keep='first')
    
    df_final = pd.merge(df_agrupado, df_srm, on='SC_ID_Key', how='inner')
    
    if 'Definição do projeto' in df_final.columns: df_final.rename(columns={'Definição do projeto': 'atuação do projeto'}, inplace=True)
    if coluna_valor_correta in df_final.columns: df_final.rename(columns={coluna_valor_correta: 'Valor Total'}, inplace=True)
    
    colunas_finais = ['atuação do projeto', 'SC ID', 'Material', 'Denominação', 'Quantidade total', 'Valor Total', 'Nº doc.de referência', 'Created On', 'SC Name', 'Next Approver', 'SC Approval status', 'Received on', 'Requester']
    colunas_presentes = [col for col in colunas_finais if col in df_final.columns]
    df_final = df_final[colunas_presentes]
    
    st.success("✅ `Planilhas.py` executado!")
    return df_final


def executar_lancamento_fim_py(df_lancamento, arquivo_lcp, arquivo_resumo):
    """
    Contém a lógica EXATA do seu LançamentoFIM.py.
    As únicas mudanças são na leitura e salvamento de arquivos.
    """
    st.write("▶️ Etapa 2: Processando `LançamentoFIM.py`...")
    
    df_lcp = pd.read_excel(arquivo_lcp, sheet_name='Capex', header=3, dtype={'WBS': str})
    
    # --- Início da sua lógica original ---
    df_lcp.columns = df_lcp.columns.str.strip()
    if 'columns' in df_lcp and df_lcp.columns.has_duplicates: df_lcp = df_lcp.loc[:, ~df_lcp.columns.duplicated()]
    if not df_lancamento.empty:
        df_lancamento.dropna(subset=['SC ID', 'atuação do projeto'], inplace=True)
        df_lancamento['SC ID'] = df_lancamento['SC ID'].str.strip()
        df_lancamento['atuação do projeto'] = df_lancamento['atuação do projeto'].str.strip()
        df_lancamento = df_lancamento[df_lancamento['SC ID'] != '']
    
    if 'WBS' in df_lcp.columns: df_lcp['WBS'] = df_lcp['WBS'].str.strip()
    df_lcp_essencial = df_lcp[['WBS', 'PROJECT NAME']].drop_duplicates(subset=['WBS'])
    df_lancamento_enriquecido = pd.merge(df_lancamento, df_lcp_essencial, left_on='atuação do projeto', right_on='WBS', how='left')
    
    if not df_lancamento_enriquecido.empty:
        chaves_de_agrupamento = ['SC ID', 'atuação do projeto']
        df_agrupado = df_lancamento_enriquecido.groupby(chaves_de_agrupamento).agg({'Denominação': lambda x: '\n'.join(x.dropna().astype(str).unique()),'SC Name': 'first', 'Created On': 'first', 'Requester': 'first', 'Valor Total': 'first', 'Next Approver': 'first', 'Received on': 'first','PROJECT NAME': 'first'}).reset_index()
    else:
        df_agrupado = pd.DataFrame()
        
    mapa_colunas = {'SC ID': 'SC', 'atuação do projeto': 'WBS', 'SC Name': 'DESCRIÇÃO','Denominação': 'CONTEÚDO', 'Created On': 'DATA CRIAÇÃO', 'Requester': 'REQUISITANTE','Valor Total': 'VALOR', 'Next Approver': 'PENDENTE COM','Received on': 'RECEBIDA EM', 'PROJECT NAME': 'PROJETO'}
    df_atualizacao = df_agrupado.rename(columns=mapa_colunas)
    
    if not df_atualizacao.empty:
        df_atualizacao['SC'] = pd.to_numeric(df_atualizacao['SC'], errors='coerce').astype('Int64').astype(str)
        df_atualizacao = df_atualizacao[df_atualizacao['SC'] != '<NA>']
        
    workbook = load_workbook(arquivo_resumo)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    col_map = {name: i+1 for i, name in enumerate(headers)}
    sc_col_num = col_map.get('SC'); wbs_col_num = col_map.get('WBS')
    if not sc_col_num or not wbs_col_num: raise ValueError("Colunas 'SC' e 'WBS' devem existir.")
    key_row_map = {}
    for i in range(2, sheet.max_row + 1):
        sc_val = str(sheet.cell(row=i, column=sc_col_num).value or '').replace('.0', '').strip()
        wbs_val = str(sheet.cell(row=i, column=wbs_col_num).value or '').strip()
        if sc_val and wbs_val: key_row_map[(sc_val, wbs_val)] = i
        
    colunas_gerenciadas = list(mapa_colunas.values())
    for _, row_data in df_atualizacao.iterrows():
        sc_id = str(row_data.get('SC')); wbs_id = str(row_data.get('WBS'))
        chave_unica = (sc_id, wbs_id)
        if chave_unica in key_row_map:
            target_row = key_row_map[chave_unica]
            for col_name in colunas_gerenciadas:
                if col_name in col_map and col_name in row_data.index:
                    target_col = col_map[col_name]
                    sheet.cell(row=target_row, column=target_col).value = row_data[col_name]
        else:
            new_row_values = [row_data.get(header_name) for header_name in headers]
            sheet.append(new_row_values)
            
    # Formatação final
    header_font = Font(bold=True, color="FFFFFF"); header_fill = PatternFill(start_color="002D62", end_color="002D62", fill_type="solid"); borda_fina = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')); alinhamento_central_sem_quebra = Alignment(horizontal='center', vertical='center'); alinhamento_central_com_quebra = Alignment(horizontal='center', vertical='center', wrap_text=True)
    colunas_com_quebra = ['DESCRIÇÃO', 'CONTEÚDO']
    sheet.row_dimensions[1].height = 25
    for cell in sheet[1]: cell.font = header_font; cell.fill = header_fill; cell.border = borda_fina; cell.alignment = alinhamento_central_sem_quebra
    last_data_row = sheet.max_row
    for row in sheet.iter_rows(min_row=2, max_row=last_data_row):
        for cell in row:
            cell.border = borda_fina; header_da_celula = sheet.cell(row=1, column=cell.column).value
            if header_da_celula in colunas_com_quebra: cell.alignment = alinhamento_central_com_quebra
            else: cell.alignment = alinhamento_central_sem_quebra
    if 'VALOR' in col_map:
        letra_col_valor = chr(ord('A') + col_map['VALOR'] - 1)
        for cell in sheet[letra_col_valor][1:]:
            if isinstance(cell.value, (int, float)): cell.number_format = 'R$ #,##0.00'
    larguras = {'SC': 15, 'WBS': 25, 'PROJETO': 45, 'DESCRIÇÃO': 45, 'CONTEÚDO': 50, 'VALOR': 18, 'DATA CRIAÇÃO': 18, 'REQUISITANTE': 25, 'RECEBIDA EM': 18, 'PENDENTE COM': 25, 'STATUS': 15, 'OK': 10, 'COMENTARIO': 50, 'Complemento dos materiais': 50}
    for col_name, width in larguras.items():
        if col_name in col_map:
            letra_col = chr(ord('A') + col_map[col_name] - 1)
            sheet.column_dimensions[letra_col].width = width
            
    # Prepara o arquivo para download
    virtual_workbook = io.BytesIO()
    workbook.save(virtual_workbook)
    st.success("✅ `LançamentoFIM.py` executado!")
    return virtual_workbook.getvalue()

# --- 3. INTERFACE DO APLICATIVO ---
st.title("🤖 Ferramenta de Automação de Lançamentos - FollowUP GY")
st.markdown("---")
st.header("1. Carregue os 4 arquivos necessários")

col1, col2 = st.columns(2)
with col1:
    upload_gestao = st.file_uploader("1. Planilha de Gestão (a ser atualizada)", type="xlsx")
    upload_cji5 = st.file_uploader("2. `resultado_cji5.xlsx`", type="xlsx")
with col2:
    upload_srm = st.file_uploader("3. `DADOS_SRM.xlsx`", type="xlsx")
    upload_lcp = st.file_uploader("4. `BUSCAR_LCP.xlsx`", type="xlsx")

st.markdown("---")

if upload_gestao and upload_cji5 and upload_srm and upload_lcp:
    st.header("2. Execute a Automação Completa")
    if st.button("🚀 Gerar Relatório Final Atualizado"):
        with st.spinner("Etapa 1: Executando lógica de `Planilhas.py`..."):
            df_intermediario = executar_planilhas_py(upload_cji5, upload_srm)
        
        if not df_intermediario.empty:
            with st.spinner("Etapa 2: Executando lógica de `LançamentoFIM.py`..."):
                dados_finais_para_download = executar_lancamento_fim_py(df_intermediario, upload_lcp, upload_gestao)
                
                st.success("🎉 Processo Concluído com Sucesso!")
                st.balloons()
                
                st.download_button(
                    label="📥 Baixar Planilha de Gestão FINAL",
                    data=dados_finais_para_download,
                    file_name="Gestão_de_SC_em_aberto_ATUALIZADA.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.warning("A primeira etapa não gerou dados para lançamento. Verifique os arquivos de entrada.")
else:
    st.info("Por favor, carregue todos os 4 arquivos para habilitar o botão de processamento.")
